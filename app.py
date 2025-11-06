#!/usr/bin/env python3
"""
CLI Application para Sales Forecast
Usa todos los modelos avanzados del model-service
"""
import sys
import os
import json
from pathlib import Path
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import requests

# ANSI colors para terminal
class Colors:
    HEADER = '\033[95m'
    BLUE = '\033[94m'
    CYAN = '\033[96m'
    GREEN = '\033[92m'
    YELLOW = '\033[93m'
    RED = '\033[91m'
    END = '\033[0m'
    BOLD = '\033[1m'

def print_header():
    """Imprimir encabezado bonito"""
    print(f"""
{Colors.CYAN}╔══════════════════════════════════════════════════════════════╗
║                                                              ║
║        📊 SALES FORECAST PRO - CLI Application 📊            ║
║                                                              ║
║    Modelos: Prophet | SARIMAX | Ensemble | ML Cluster       ║
║                                                              ║
╚══════════════════════════════════════════════════════════════╝{Colors.END}
""")

def parse_excel(file_path):
    """
    Parsear archivo Excel y convertir a formato requerido por la API
    """
    print(f"{Colors.BLUE}📂 Leyendo archivo Excel: {file_path}{Colors.END}")
    
    try:
        # Leer Excel
        df = pd.read_excel(file_path, header=None)
        
        concept_row = df.iloc[0].tolist() if len(df) > 0 else []
        date_row = df.iloc[1].tolist() if len(df) > 1 else []
        
        data_start_col = 9
        
        meses = []
        meses_fecha = []
        conceptos = {}
        
        for i in range(data_start_col, len(date_row)):
            date_str = str(date_row[i]).strip()
            concepto = str(concept_row[i]).strip() if i < len(concept_row) else ""
            
            if not date_str or date_str == 'nan':
                continue
            
            try:
                if isinstance(date_row[i], (int, float)):
                    fecha = pd.to_datetime('1899-12-30') + pd.Timedelta(days=date_row[i])
                else:
                    fecha = pd.to_datetime(date_str)
                
                mes_key = fecha.strftime("%Y-%m")
                meses.append(mes_key)
                meses_fecha.append(fecha.strftime("%Y-%m-%d"))
                conceptos[mes_key] = concepto
                
            except:
                continue
        
        if not meses:
            raise ValueError("No se encontraron fechas válidas en el Excel")
        
        print(f"{Colors.GREEN}✅ Fechas: {len(meses)} meses ({meses[0]} a {meses[-1]}){Colors.END}")
        
        clientes_dict = {}
        total_skus = 0
        
        for row_idx in range(2, len(df)):
            row = df.iloc[row_idx].tolist()
            
            codigo_completo = str(row[0]).strip() if len(row) > 0 else ""
            
            if not codigo_completo or len(codigo_completo) < 35:
                continue
            
            codigo_cliente = codigo_completo[:28]
            codigo_articulo = codigo_completo[-7:]
            
            ventas_mes = {}
            for i, mes in enumerate(meses):
                col_idx = data_start_col + i
                if col_idx < len(row):
                    val = row[col_idx]
                    ventas_mes[mes] = float(val) if pd.notna(val) else 0.0
                else:
                    ventas_mes[mes] = 0.0
            
            sku_id = f"{codigo_cliente}-{codigo_articulo}"
            sku_data = {
                "skuId": sku_id,
                "codigoCliente": codigo_cliente,
                "codigoArticulo": codigo_articulo,
                "grupoMatDesc": str(row[4]) if len(row) > 4 else "",
                "marcaDesc": str(row[5]) if len(row) > 5 else "",
                "negocioDesc": str(row[8]) if len(row) > 8 else "",
                "ventasMes": ventas_mes
            }
            
            if codigo_cliente not in clientes_dict:
                clientes_dict[codigo_cliente] = {
                    "codigo": codigo_cliente,
                    "skus": []
                }
            
            clientes_dict[codigo_cliente]["skus"].append(sku_data)
            total_skus += 1
        
        clientes = list(clientes_dict.values())
        
        print(f"{Colors.GREEN}✅ Procesados: {len(clientes)} clientes, {total_skus} SKUs{Colors.END}")
        
        return {
            "clientes": clientes,
            "meses": meses,
            "mesesFecha": meses_fecha,
            "conceptos": conceptos,
            "totalSkus": total_skus
        }
        
    except Exception as e:
        print(f"{Colors.RED}❌ Error al leer Excel: {str(e)}{Colors.END}")
        sys.exit(1)

def get_user_input():
    """Obtener parámetros del usuario"""
    print(f"\n{Colors.BOLD}📋 Configuración del Forecast{Colors.END}\n")
    
    while True:
        file_path = input(f"{Colors.CYAN}📁 Ruta del archivo Excel: {Colors.END}").strip()
        if os.path.exists(file_path):
            break
        print(f"{Colors.RED}❌ Archivo no encontrado{Colors.END}")
    
    sales_data = parse_excel(file_path)
    meses = sales_data["meses"]
    
    print(f"\n{Colors.CYAN}📅 Últimos meses disponibles:{Colors.END}")
    for mes in meses[-6:]:
        print(f"   • {mes}")
    
    print(f"\n{Colors.CYAN}📆 Fecha de corte (último mes histórico):{Colors.END}")
    end_date = input(f"   Mes (YYYY-MM) [{meses[-1]}]: ").strip() or meses[-1]
    
    while True:
        try:
            forecast_months = input(f"{Colors.CYAN}🔮 Meses a pronosticar [6]: {Colors.END}").strip()
            forecast_months = int(forecast_months) if forecast_months else 6
            if 1 <= forecast_months <= 24:
                break
        except ValueError:
            pass
        print(f"{Colors.RED}❌ Debe ser entre 1 y 24{Colors.END}")
    
    print(f"\n{Colors.CYAN}🧠 Modelo de forecasting:{Colors.END}")
    print("   1. Prophet (recomendado)")
    print("   2. SARIMAX")
    print("   3. Holt-Winters")
    print("   4. Ensemble (más preciso)")
    print("   5. ML Cluster")
    
    modelos = {"1": "prophet", "2": "sarimax", "3": "holtwinters", "4": "ensemble", "5": "ml_cluster"}
    modelo_opt = input(f"{Colors.CYAN}Selecciona [4]: {Colors.END}").strip() or "4"
    modelo = modelos.get(modelo_opt, "ensemble")
    
    return {
        "file_path": file_path,
        "sales_data": sales_data,
        "end_date": end_date,
        "forecast_months": forecast_months,
        "model": modelo,
        "options": {
            "seasonal_period": 12,
            "n_clusters": 5,
            "enable_survival": True,
            "cv_folds": 3
        }
    }

def call_forecast_api(payload, model_service_url="http://localhost:8000"):
    """Llamar a la API del model service"""
    print(f"\n{Colors.BLUE}🚀 Ejecutando modelo {payload['model'].upper()}...{Colors.END}")
    print(f"{Colors.YELLOW}⏳ Esto puede tardar 30-60 segundos...{Colors.END}")
    
    try:
        response = requests.post(
            f"{model_service_url}/predict",
            json=payload,
            timeout=300
        )
        
        if response.status_code == 200:
            print(f"{Colors.GREEN}✅ Forecast completado{Colors.END}")
            return response.json()
        else:
            error_data = response.json()
            print(f"{Colors.RED}❌ Error: {error_data.get('detail')}{Colors.END}")
            return None
            
    except requests.exceptions.ConnectionError:
        print(f"{Colors.RED}❌ No se pudo conectar al servicio en {model_service_url}{Colors.END}")
        print(f"{Colors.YELLOW}💡 Ejecuta: docker-compose up{Colors.END}")
        return None
    except Exception as e:
        print(f"{Colors.RED}❌ Error: {str(e)}{Colors.END}")
        return None

def create_excel_report(result, output_path):
    """Crear Excel con resultados"""
    print(f"\n{Colors.BLUE}📝 Creando Excel...{Colors.END}")
    
    wb = openpyxl.Workbook()
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Resumen
    ws = wb.active
    ws.title = "Resumen"
    ws.append(["SALES FORECAST - RESUMEN EJECUTIVO"])
    ws.append([])
    ws.append(["Métrica", "Valor"])
    ws.append(["Modelo", result["modelUsed"]])
    ws.append(["Ventas Históricas €", f"{result['summary']['totalVentasHistoricas']:,.2f}"])
    ws.append(["Forecast Total €", f"{result['summary']['totalForecast']:,.2f}"])
    ws.append(["Crecimiento %", f"{result['summary']['crecimientoEsperado']:.2f}"])
    ws.append(["Clientes Activos", result['summary']['clientesActivos']])
    ws.append(["Tiempo (s)", result['diagnostics']['training_time_s']])
    
    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Forecast por SKU
    ws_sku = wb.create_sheet("Forecast por SKU")
    headers = ["Cliente", "SKU", "Última Venta €", "Forecast €", "Variación %", "Estado", "MAPE %"]
    forecast_meses = result["detalleCompleto"]["forecastMeses"]
    for mes in forecast_meses:
        headers.append(f"Forecast {mes}")
    
    ws_sku.append(headers)
    for cell in ws_sku[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    for sku in result["detalleCompleto"]["resultadosPorSku"]:
        row = [
            sku["codigoCliente"],
            sku["codigoArticulo"],
            sku["ultimoValor"],
            sku["forecast"],
            sku["variacion"],
            sku["estado"],
            sku.get("mape", "")
        ]
        for val in sku.get("forecast_detalle", []):
            row.append(val)
        ws_sku.append(row)
    
    # Forecast por Cliente
    ws_cli = wb.create_sheet("Forecast por Cliente")
    ws_cli.append(["Cliente", "Ventas Históricas €", "Forecast €", "Variación %", "SKUs Activos"])
    for cell in ws_cli[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    for cli in result["detalleCompleto"]["resultadosPorCliente"]:
        ws_cli.append([
            cli["codigoCliente"],
            cli["ventasHistorico"],
            cli["ventasForecast"],
            cli["variacion"],
            cli["skusActivos"]
        ])
    
    # Ajustar anchos
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
    
    wb.save(output_path)
    print(f"{Colors.GREEN}✅ Excel guardado: {output_path}{Colors.END}")

def main():
    print_header()
    
    params = get_user_input()
    
    payload = {
        "salesData": params["sales_data"],
        "forecastMonths": params["forecast_months"],
        "endDate": params["end_date"],
        "model": params["model"],
        "options": params["options"]
    }
    
    result = call_forecast_api(payload)
    
    if not result:
        sys.exit(1)
    
    print(f"\n{Colors.GREEN}{Colors.BOLD}📊 RESULTADOS{Colors.END}")
    print(f"{Colors.CYAN}{'='*50}{Colors.END}")
    print(f"Modelo: {Colors.BOLD}{result['modelUsed']}{Colors.END}")
    print(f"Forecast: {Colors.BOLD}€{result['summary']['totalForecast']:,.2f}{Colors.END}")
    print(f"Crecimiento: {Colors.BOLD}{result['summary']['crecimientoEsperado']:.2f}%{Colors.END}")
    print(f"Tiempo: {Colors.BOLD}{result['diagnostics']['training_time_s']:.2f}s{Colors.END}")
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"forecast_{result['modelRequested']}_{timestamp}.xlsx"
    create_excel_report(result, output_file)
    
    print(f"\n{Colors.GREEN}{Colors.BOLD}🎉 ¡Completado!{Colors.END}")
    print(f"{Colors.CYAN}📁 {output_file}{Colors.END}\n")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print(f"\n{Colors.YELLOW}⚠️  Cancelado{Colors.END}")
        sys.exit(0)
    except Exception as e:
        print(f"\n{Colors.RED}❌ Error: {str(e)}{Colors.END}")
        sys.exit(1)
